"""
Export module for MPDTE College Predictor
Handles CSV, Excel, and PDF report generation
"""
import os
import csv
import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)
EXPORT_DIR = os.path.join(Path.home(), "mpdte_exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def export_csv(data: list[dict], filename: str = None) -> str:
    if not data:
        return ""
    if not filename:
        filename = f"mpdte_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(EXPORT_DIR, filename)
    keys = list(data[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in keys})
    return filepath


def export_excel(data: list[dict], filename: str = None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not available")
        return ""

    if not data:
        return ""
    if not filename:
        filename = f"mpdte_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MPDTE Predictions"

    # Header style
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    keys = [
        "institute_name", "branch_full_name", "institute_type",
        "opening_rank", "closing_rank", "allotted_category",
        "probability", "risk_level", "fee_waiver_available",
        "domicile", "round_info", "year", "city"
    ]
    friendly = [
        "College Name", "Branch", "Type",
        "Opening Rank", "Closing Rank", "Category",
        "Probability %", "Risk Level", "Fee Waiver",
        "Domicile", "Round", "Year", "City"
    ]

    for col_idx, header in enumerate(friendly, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    prob_colors = {
        "Almost Guaranteed": "00C851",
        "Very Safe": "7CB342",
        "Safe": "FFD600",
        "Possible": "FF6D00",
        "Difficult": "D32F2F",
        "Very Unlikely": "880E4F",
    }

    for row_idx, rec in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            val = rec.get(key, "")
            if key == "fee_waiver_available":
                val = "Yes" if val else "No"
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Color by risk
        risk = rec.get("risk_level", "")
        color = prob_colors.get(risk, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col_idx in range(1, len(keys) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-width
    for col_idx, key in enumerate(keys, 1):
        max_len = len(friendly[col_idx - 1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filepath)
    return filepath


def export_pdf_report(summary: dict, filename: str = None) -> str:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
    except ImportError:
        logger.error("reportlab not available")
        return ""

    if not filename:
        filename = f"mpdte_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(EXPORT_DIR, filename)

    doc = SimpleDocTemplate(
        filepath, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#0f3460"),
        spaceAfter=12
    )
    heading_style = ParagraphStyle(
        "Heading", parent=styles["Heading2"],
        fontSize=12, textColor=colors.HexColor("#16213e"),
        spaceAfter=6
    )
    normal_style = ParagraphStyle(
        "Normal", parent=styles["Normal"],
        fontSize=9, spaceAfter=4
    )

    story = []

    # Title
    story.append(Paragraph("MPDTE College Predictor - Admission Analysis Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", normal_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#0f3460")))
    story.append(Spacer(1, 0.3*cm))

    # Summary box
    story.append(Paragraph("Your Profile", heading_style))
    profile_data = [
        ["JEE Rank", str(summary.get("user_rank", "N/A")),
         "Category", summary.get("category", "N/A")],
        ["MP Domicile", "Yes" if summary.get("domicile") == "Yes" else "No",
         "Fee Waiver", "Yes" if summary.get("fee_waiver") else "No"],
        ["Eligible Colleges", str(summary.get("total_eligible_colleges", 0)),
         "Eligible Branches", str(summary.get("total_eligible_branches", 0))],
    ]
    t = Table(profile_data, colWidths=[3*cm, 4*cm, 3*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f4ff")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#e8eeff"), colors.HexColor("#f5f5f5")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # Probability Summary
    prob_sum = summary.get("probability_summary", {})
    story.append(Paragraph("Admission Probability Summary", heading_style))
    prob_data = [
        ["Category", "Count"],
        ["✅ Almost Guaranteed (95%+)", str(prob_sum.get("guaranteed", 0))],
        ["🟢 Very Safe (80-95%)", str(prob_sum.get("very_safe", 0))],
        ["🟡 Safe (60-80%)", str(prob_sum.get("safe", 0))],
        ["🟠 Possible (40-60%)", str(prob_sum.get("possible", 0))],
        ["🔴 Difficult (20-40%)", str(prob_sum.get("difficult", 0))],
        ["⚫ Very Unlikely (<20%)", str(prob_sum.get("unlikely", 0))],
    ]
    t2 = Table(prob_data, colWidths=[8*cm, 3*cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#e8f5e9"), colors.HexColor("#f5f5f5")]),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.4*cm))

    # Top 20 recommendations
    story.append(Paragraph("Top 20 Recommended Colleges", heading_style))
    top20 = summary.get("top_20", [])

    table_header = ["#", "College Name", "Branch", "Type", "Open Rank", "Close Rank", "Category", "Prob%", "Risk"]
    table_data = [table_header]

    risk_colors_map = {
        "Almost Guaranteed": colors.HexColor("#00C851"),
        "Very Safe": colors.HexColor("#7CB342"),
        "Safe": colors.HexColor("#FFD600"),
        "Possible": colors.HexColor("#FF6D00"),
        "Difficult": colors.HexColor("#D32F2F"),
        "Very Unlikely": colors.HexColor("#880E4F"),
    }

    for i, rec in enumerate(top20[:20], 1):
        college = rec.get("institute_name", "")[:35]
        branch = rec.get("branch_full_name", "")[:30]
        table_data.append([
            str(i),
            Paragraph(college, normal_style),
            Paragraph(branch, normal_style),
            rec.get("institute_type", "")[:8],
            f"{rec.get('opening_rank', 0):,}" if rec.get("opening_rank") else "-",
            f"{rec.get('closing_rank', 0):,}" if rec.get("closing_rank") else "-",
            rec.get("allotted_category", "")[:15],
            f"{rec.get('probability', 0):.0f}%",
            rec.get("risk_level", "")[:12],
        ])

    col_widths = [0.6*cm, 7*cm, 6.5*cm, 1.5*cm, 2.2*cm, 2.2*cm, 3*cm, 1.5*cm, 3*cm]
    t3 = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f9f9f9"), colors.white]),
    ]
    t3.setStyle(TableStyle(style_cmds))
    story.append(t3)
    story.append(Spacer(1, 0.3*cm))

    # Best options summary
    story.append(Paragraph("Best Options Summary", heading_style))
    bests = []
    for label, key in [
        ("Best CS Option", "best_cs_option"),
        ("Best Core Engineering", "best_core_option"),
        ("Best Government College", "best_govt_option"),
        ("Best Private College", "best_private_option"),
        ("Best Fee Waiver Option", "best_fw_option"),
    ]:
        rec = summary.get(key)
        if rec:
            bests.append([
                label,
                rec.get("institute_name", "")[:40],
                rec.get("branch_full_name", "")[:30],
                f"{rec.get('probability', 0):.0f}%",
            ])

    if bests:
        best_header = [["Category", "College", "Branch", "Probability"]]
        t4 = Table(best_header + bests, colWidths=[5*cm, 9*cm, 7*cm, 3*cm])
        t4.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e94560")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#fff3f3"), colors.white]),
        ]))
        story.append(t4)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "Disclaimer: This report is based on historical MPDTE counselling data. "
        "Actual admissions may vary based on seat availability, updated cutoffs, and counselling policy changes.",
        ParagraphStyle("Disclaimer", parent=normal_style, textColor=colors.grey, fontSize=8)
    ))

    doc.build(story)
    return filepath
